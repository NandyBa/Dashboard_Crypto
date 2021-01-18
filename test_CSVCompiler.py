from unittest import TestCase, mock
import unittest
import CSVCompiler
import os
import openpyxl

class TestCSVCompiler(TestCase):

    def compilecsv(self, exchanges, file = "summary_crypto.xlsx"):
        CSVCompiler.CSVCompiler().compilecsv(exchanges, file)

    @mock.patch('CoinbaseExchange.CoinbaseExchange.get_balances', return_value = [
        {"amount":3, "currency":"BTC"},
        {"amount":2.45553, "currency":"ETH"}
    ])
    def test_Coinbase_compile(self, mock_get_balances):
        file = "testcompiler.xlsx"
        self.compilecsv(["Coinbase"], file)

        wb_obj = openpyxl.load_workbook(file)
        sheet = wb_obj['Balances']
        i = 0
        count = 0
        for row in sheet.iter_rows():
            if i == 0:
                finded = False
                for j in range(len(row)):
                    if row[j].value == "Coinbase":
                        iCoinbaseAmount = j
                        count += 1
                if count != 1:
                    raise ValueError("Error in the Coinbase compiler output")
            else:
                if(row[0].value == "BTC"):
                    self.assertEqual(float(row[iCoinbaseAmount].value), 3)
                    count += 1
                elif(row[0].value == "ETH"):
                    self.assertEqual(float(row[iCoinbaseAmount].value), 2.45553)
                    count += 1
            i += 1
        self.assertEqual(count, 3)
        os.remove(file)

    @mock.patch('BinanceExchange.BinanceExchange.get_balances', return_value = [
        {'asset': 'BTC', 'free': 0.00000856, 'locked': 0.00000013},
        {'asset': 'LTC', 'free': 0.00000000, 'locked': 0.17000000},
        {'asset': 'LDETH', 'free': 0.56000000, 'locked': 0.00000000},
        {'asset': 'LDEOS', 'free': 3.00000000, 'locked': 0.00000000}
    ])
    def test_Binance_compile(self, mock_get_balances):
        file = "testcompiler2.xlsx"
        self.compilecsv(["Binance"], file)

        wb_obj = openpyxl.load_workbook(file)
        sheet = wb_obj['Balances']

        i = 0
        count = 0
        for row in sheet.iter_rows():
            if i == 0:
                for j in range(len(row)):
                    if row[j].value == "Binance":
                        iBinanceAmount = j
                        count += 1
                    elif row[j].value == "Binance Flexible Saving":
                        iBinanceAmountFlexibleSaving = j
                        count += 1
                if count != 2:
                    raise ValueError("Error in the Binance compiler output")
            else:
                if(row[0].value == "BTC"):
                    self.assertEqual(float(row[iBinanceAmount].value), 8.69e-06)
                    count += 1
                elif(row[0].value == "LTC"):
                    self.assertEqual(float(row[iBinanceAmount].value), 0.17)
                    count += 1
                elif(row[0].value == "ETH"):
                    self.assertEqual(float(row[iBinanceAmountFlexibleSaving].value), 0.56)
                    count += 1
                elif(row[0].value == "EOS"):
                    self.assertEqual(float(row[iBinanceAmountFlexibleSaving].value), 3)
                    count += 1
            i += 1
        self.assertEqual(count, 6)
        os.remove(file)


    @mock.patch('CoinbaseExchange.CoinbaseExchange.get_balances', return_value = [
        {"amount":2.5, "currency":"BTC"},
        {"amount":2, "currency":"ETH"}
    ])
    @mock.patch('BinanceExchange.BinanceExchange.get_balances', return_value = [
        {'asset': 'BTC', 'free': 0.5, 'locked': 0},
        {'asset': 'LTC', 'free': 2, 'locked': 0},
        {'asset': 'LDETH', 'free': 1, 'locked': 1},
        {'asset': 'LDEOS', 'free': 3, 'locked': 0}
    ])
    @mock.patch('Exchanges.Exchanges.get_Price', side_effect=[
        30000,      #BTC
        5,          #EOS
        1000,       #ETH
        120,        #LTC
    ])
    # Total 
    # 3 BTC => 90,000
    # 3 EOS => 15
    # 4 ETH => 4,000
    # 2 LTC => 240
    #
    def test_compile_financial_statement(self, mock_Coinbase_get_balances, mock_Binance_get_balances, mock_Exchanges_get_Price):
        file = "testcompiler3.xlsx"
        csvCompiler = CSVCompiler.CSVCompiler()
        csvCompiler.compilecsv(["Coinbase","Binance"], file)

        csvCompiler.compile_financial_statement(file)

        wb_obj = openpyxl.load_workbook(file)
        sheetToRead = wb_obj['Financial Statement']

        number_of_test_passed = 0

        for row in sheetToRead.iter_rows():
            if row[1].value == "BTC" :
                self.assertEqual(row[2].value, 30000) # Price for 1 asset
                self.assertEqual(row[3].value, 90000) # Total for user bag
                number_of_test_passed += 2

            elif row[1].value == "ETH" :
                self.assertEqual(row[2].value, 1000) # Price for 1 asset
                self.assertEqual(row[3].value, 4000) # Total for user bag
                number_of_test_passed += 2

            elif row[1].value == "LTC" :
                self.assertEqual(row[2].value, 120) # Price for 1 asset
                self.assertEqual(row[3].value, 240) # Total for user bag
                number_of_test_passed += 2

            elif row[1].value == "EOS" :
                self.assertEqual(row[2].value, 5) # Price for 1 asset
                self.assertEqual(row[3].value, 15) # Total for user bag
                number_of_test_passed += 2

        self.assertEqual(number_of_test_passed, 8)
        os.remove(file)

if __name__ == '__main__':
    unittest.main()