import pandas as pd
import CoinbaseExchange
import BinanceExchange
import Exchanges
import xlsxwriter
import openpyxl

class CSVCompiler():
    def compilecsv(self, exchanges = [], file = "summary_crypto.xlsx"):
        Coinbase = pd.Series(dtype='int', name="Coinbase")
        Binance = pd.Series(dtype='int', name="Binance")
        BinanceFlexibleSaving = pd.Series(dtype='int', name="BinanceEarn")
        if("Coinbase" in exchanges):
            c = CoinbaseExchange.CoinbaseExchange()
            CoinbaseBalances  = c.get_balances()
            
            for balance in CoinbaseBalances:
                Coinbase[balance["currency"]] = balance["amount"]
        if("Binance" in exchanges):
            c = BinanceExchange.BinanceExchange()
            BinanceBalances  = c.get_balances()
            
            for balance in BinanceBalances:
                if "LD" in balance["asset"][0:2]:
                    BinanceFlexibleSaving[balance["asset"].replace("LD","")] = balance["free"] + balance["locked"]
                else:
                    Binance[balance["asset"]] = balance["free"] + balance["locked"]
        
        
        df = pd.DataFrame({ "Coinbase": Coinbase,"Binance": Binance, "Binance Flexible Saving": BinanceFlexibleSaving})
        writer = pd.ExcelWriter(file, engine='xlsxwriter')
        df.to_excel(writer, sheet_name="Balances")
        writer.save()

    def compile_financial_statement(self, file = "summary_crypto.xlsx"):
        wb_obj = openpyxl.load_workbook(file)
        sheetToRead = wb_obj['Balances']
        ex = Exchanges.Exchanges()

        financial_statement_sheet = wb_obj.create_sheet(title="Financial Statement")

        Prices = pd.Series(dtype='float', name="Price")
        Valorisations = pd.Series(dtype='int', name="Valorisation")

        for row in sheetToRead.iter_rows():
            if row[0].value != None :
                asset = row[0].value
                Prices[asset] = ex.get_Price(asset,"EUR")
                sum_ = 0
                for i in range(1,4):
                    if row[i].value == None:
                        sum_ += 0
                    else:
                        sum_ += row[i].value

                Valorisations[asset] = Prices[asset] * sum_

        _ = financial_statement_sheet.cell(column=2, row=1, value="Abreviation")
        _ = financial_statement_sheet.cell(column=3, row=1, value="EUR/U")
        _ = financial_statement_sheet.cell(column=4, row=1, value="")
        for i in range(2,len(Prices)+2):
            financial_statement_sheet.cell(column=1, row=i, value=Valorisations[i-2]/sum(Valorisations)).number_format = '0.00%'
            _ = financial_statement_sheet.cell(column=2, row=i, value=Prices.index[i-2])
            _ = financial_statement_sheet.cell(column=3, row=i, value=Prices[i-2])
            _ = financial_statement_sheet.cell(column=4, row=i, value=Valorisations[i-2])


        _ = financial_statement_sheet.cell(column=1, row=len(Prices)+3, value="Total:")
        _ = financial_statement_sheet.cell(column=4, row=len(Prices)+3, value=sum(Valorisations))
        #wb_obj._sheets.sort(key=lambda ws: ws.title)

        wb_obj.save(file)
        #Make Financial Statement Final in first position
        fs_sheet_actual_position = wb_obj.worksheets.index(financial_statement_sheet) #get position of new sheet financial statement sheet
        fs_sheet_new_position = 0
        sheets = wb_obj._sheets.copy()
        sheets.insert(fs_sheet_new_position, sheets.pop(fs_sheet_actual_position)) #modifying the sheets list
        wb_obj._sheets = sheets
        wb_obj.save(file)
